from openpyxl import load_workbook
from unidecode import unidecode
import operator
import asyncio
import logging

def set_logger(name):
	global logger
	logger = logging.getLogger(name+'.ss_parser')

HEADERS = ['CLIENTE', 'CONCEITO', 'NIF', 'DATA EMISSÃO', 'MONTANTE COM IVA', 'NR. DOC.', 'MORADA']

def tokenize(s, exclude_terms=[]):
	if s is None or s == '':
		return None
	exclude_terms+= [' ', '.']
	s = unidecode(s).strip()
	for term in exclude_terms:
		s = s.replace(term, '')
	return s.lower()


class ss_001(object):
	def __init__(self, name, callback, update_callback, **kwargs):
		self._callback = callback
		self._update_callback = update_callback
		set_logger(name)
		self._loop = asyncio.get_event_loop()

	def _process(self, filename):
		logger.debug('_process called.')
		logger.debug('File: %s'%filename)

		try:
			wb = load_workbook(filename, read_only=True)
			sheets = [wb.get_sheet_by_name(x) for x in wb.get_sheet_names()]
			logger.debug('File opened')
			headers = {}
			header_index = None

			for sheet in sheets:
				for row_index, row in enumerate(sheet.iter_rows()):
					for cell_index, cell in enumerate(list(row)):
						token = tokenize(cell.value)
						if token is None:
							continue
						headers[token] = cell_index
					if headers != {}:
						header_index = row_index
						break

			c_headers = [tokenize(s) for s in HEADERS]

			data = []
			for sheet in sheets:
				for row in sheet.iter_rows(row_offset=header_index+1):
					l = []
					for key in c_headers:
						l.append(row[headers[key]].value)
					data.append(l)
			logger.debug('Data aquired.')
			return data
		except:
			logger.debug('Error opening file.')
			return None

	async def process(self, filename):
		logger.debug('Processing..')
		count = 0
		data = await self._loop.run_in_executor(None, self._process, filename)
		processed = {}
		logger.debug('Data acquired.')
		
		if data is None:
			logger.debug('No data obtained. Invalid Format.')
			return count, "invalid_format"

		cleaned_data = await self.process_data(data)
		for x in cleaned_data:
			logger.debug('Data x processed.')
			if await self._callback(*x):
				logger.debug('Count + 1')
				count +=1

		await self._update_callback()
		logger.debug('All processed. %s data entries.'%count)
		return count, None

	async def process_data(self, data):
		cleaned = []
		done = []

		for x in data:
			if x[0] in done:
				pass
			else:
				done.append(x[0])
				new = []
				new.append(x[0])
				new.append(x[2])
				new.append(x[3])
				val2 = await self.get_value(data, x[1], x[5])
				total = val2 + x[4]
				new.append(total)
				new.append(x[5])

				cleaned.append(new)
		return cleaned


	async def get_value(self, data, concept, doc):
		value = [x[4] for x in data if x[1] != concept and x[5] == doc]
		if value:
			return value[0]
		else:
			return 0