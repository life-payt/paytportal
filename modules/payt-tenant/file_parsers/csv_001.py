from unidecode import unidecode
from openpyxl import load_workbook
import operator
import csv
import asyncio
import chardet
import logging

def set_logger(name):
	global logger
	logger = logging.getLogger(name+'.csv_parser')

HEADERS = ['NumeroCliente', 'Contribuinte', 'Nome', 'Policia', 'Rua', 'Andar', 'Localidade', 'Freguesia', 'Cod.postal', 'Classe consumo', 'UltimoDiaFaturado']
EXCLUDE = ['A.']
DELIMITER = ';'

def tokenize(s, exclude_terms=[]):
	if s is None or s == '':
		return None
	exclude_terms+= [' ', '.']
	s = unidecode(s).strip()
	for term in exclude_terms:
		s = s.replace(term, '')
	return s.lower()


class csv_001(object):
	def __init__(self, name, callback, update_callback, **kwargs):
		self._callback = callback
		self._update_callback = update_callback
		set_logger(name)
		self._loop = asyncio.get_event_loop()
		self._exclusions = self._load_exclusions()

	def _load_exclusions(self):
		data = []
		try:
			f = open('/tenant/file_parsers/exclusions', 'r')
			x = f.read().splitlines()
			f.close()
		except:
			return data

		for i in x:
			data.append(i)
		return data


	def _process(self, filename):
		logger.debug('_process called.')
		logger.debug('File: %s'%filename)
		data = []

		if '.csv' not in filename:
			wb = load_workbook(filename, read_only=True)
			sheets = [wb.get_sheet_by_name(x) for x in wb.get_sheet_names()]
			logger.debug('File opened')
			headers = {}
			header_index = None

			for sheet in sheets:
				for row_index, row in enumerate(sheet.iter_rows()):
					for cell_index, cell in enumerate(list(row)):
						token = tokenize(cell.value, EXCLUDE)
						if token is None:
							continue
						headers[token] = cell_index
					if headers != {}:
						header_index = row_index
						break

			c_headers = [tokenize(s, EXCLUDE) for s in HEADERS]

			data = []
			for sheet in sheets:
				for row in sheet.iter_rows(row_offset=header_index+1):
					l = []
					for key in c_headers:
						l.append(row[headers[key]].value)
					data.append(l)
			logger.debug('Data aquired.')
			return data
		else:
			try:
				with open(filename, 'rb') as file:
					encoding = chardet.detect(file.read())['encoding']
					logger.debug('Detecting encoding format..')
			except:
				logger.debug('Exception in detection of encoding format. Set to utf-8..')
				encoding = 'utf-8'

			if encoding == None:
				encoding = 'utf-8'
			logger.debug('Encoding format found: %s'%(encoding))

			try:
				with open(filename, 'rb') as file:
					logger.debug('File opened.')
					headers = {}
					h_string = file.readline().decode(encoding).strip()
					for cell_index, cell in enumerate(h_string.split(DELIMITER)):
						token = tokenize(cell, EXCLUDE)
						if token is None:
							continue
						headers[token] = cell_index

					c_headers = [tokenize(s, EXCLUDE) for s in HEADERS]
					
					for row in file:
						row = row.decode(encoding).strip().split(DELIMITER)
						l = []
						for key in c_headers:
							l.append(row[headers[key]])

						l[-1] = int(l[-1][0])
						if l[-1] <= 2:
							data.append(l)
			except:
				logger.debug('Error opening file.')
				return None

		logger.debug('All data aquired.')
		return data

	async def process(self, filename):
		logger.debug('Processing..')
		count = 0
		data = await self._loop.run_in_executor(None, self._process, filename)
		logger.debug('Data acquired.')

		if data is None:
			logger.debug('No data obtained. Invalid Format.')
			return count, "invalid_format"

		for x in data:
			if x[8] not in self._exclusions and ("1" in x[9] or "2" in x[9]):
				if await self._callback(*x):
					logger.debug('Count + 1')
					count +=1
			else:
				logger.debug('Client excluded..')

		await self._update_callback()
		logger.debug('All processed. %s data entries.'%count)
		return count, None